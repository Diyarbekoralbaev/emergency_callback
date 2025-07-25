import json
import random

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Avg, Q
from django.utils import timezone
from datetime import datetime, timedelta
import pytz

from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from .models import CallbackRequest, Rating, CallStatus
from .forms import CallbackRequestForm
from .tasks import fixed_process_callback_call as process_callback_call
from .utils import get_message, TASHKENT_TZ
from teams.models import Team, Region
from teams.permissions import admin_required, has_role
from users.models import UserRoleChoices
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io

def convert_to_tashkent(dt):
    """Convert datetime to Tashkent timezone"""
    if not dt:
        return None
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, pytz.UTC)
    return dt.astimezone(TASHKENT_TZ)


@admin_required
def dashboard(request):
    # Get filter parameters
    region_filter = request.GET.get('region', '')
    team_filter = request.GET.get('team', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status_filter = request.GET.get('status', '')

    # Default date range (today if no filters)
    if not date_from and not date_to:
        today = timezone.now().astimezone(TASHKENT_TZ).date()
        date_from = today
        date_to = today
    else:
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = timezone.now().astimezone(TASHKENT_TZ).date() - timedelta(days=30)

        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = timezone.now().astimezone(TASHKENT_TZ).date()

    # Convert dates to UTC for database queries
    start_utc = timezone.make_aware(datetime.combine(date_from, datetime.min.time()), TASHKENT_TZ).astimezone(pytz.UTC)
    end_utc = timezone.make_aware(datetime.combine(date_to, datetime.max.time()), TASHKENT_TZ).astimezone(pytz.UTC)

    # Base querysets with date filtering
    callbacks_qs = CallbackRequest.objects.filter(
        created_at__gte=start_utc,
        created_at__lte=end_utc
    )

    ratings_qs = Rating.objects.filter(
        timestamp__gte=start_utc,
        timestamp__lte=end_utc
    )

    # Apply region filter
    if region_filter:
        callbacks_qs = callbacks_qs.filter(team__region_id=region_filter)
        ratings_qs = ratings_qs.filter(team__region_id=region_filter)

    # Apply team filter
    if team_filter:
        callbacks_qs = callbacks_qs.filter(team_id=team_filter)
        ratings_qs = ratings_qs.filter(team_id=team_filter)

    # Apply status filter
    if status_filter:
        callbacks_qs = callbacks_qs.filter(status=status_filter)

    # Call statistics
    total_calls = callbacks_qs.count()
    completed_calls = callbacks_qs.filter(
        status__in=[CallStatus.COMPLETED, CallStatus.TRANSFERRED]
    ).count()
    failed_calls = callbacks_qs.filter(status=CallStatus.FAILED).count()
    no_rating_calls = callbacks_qs.filter(status=CallStatus.NO_RATING).count()
    pending_calls = callbacks_qs.filter(
        status__in=[CallStatus.PENDING, CallStatus.DIALING, CallStatus.CONNECTING]
    ).count()

    # Rating statistics
    total_ratings = ratings_qs.count()
    avg_rating = ratings_qs.aggregate(Avg('rating'))['rating__avg'] or 0

    # Rating distribution - ensure all ratings 1-5 are shown
    rating_distribution = []
    for i in range(1, 6):
        count = ratings_qs.filter(rating=i).count()
        percentage = (count / total_ratings * 100) if total_ratings > 0 else 0
        rating_distribution.append({
            'rating': i,
            'count': count,
            'percentage': round(percentage, 1)
        })

    # Success rate calculation
    success_rate = 0
    if total_calls > 0:
        success_rate = round((completed_calls / total_calls) * 100, 1)

    # Failure rate calculation
    failure_rate = 0
    if total_calls > 0:
        failure_rate = round((failed_calls / total_calls) * 100, 1)

    # Recent calls with ratings info (respecting filters)
    recent_calls_qs = callbacks_qs.select_related('team', 'team__region', 'requested_by').prefetch_related('rating')
    recent_calls = recent_calls_qs.order_by('-created_at')[:15]

    # Daily breakdown for chart (last 7 days within filter range)
    daily_stats = []
    chart_start_date = max(date_from, timezone.now().astimezone(TASHKENT_TZ).date() - timedelta(days=6))
    chart_end_date = min(date_to, timezone.now().astimezone(TASHKENT_TZ).date())

    current_date = chart_start_date
    while current_date <= chart_end_date:
        # Convert to UTC for database queries
        day_start_utc = timezone.make_aware(datetime.combine(current_date, datetime.min.time()),
                                            TASHKENT_TZ).astimezone(pytz.UTC)
        day_end_utc = timezone.make_aware(datetime.combine(current_date, datetime.max.time()), TASHKENT_TZ).astimezone(
            pytz.UTC)

        day_calls = callbacks_qs.filter(created_at__gte=day_start_utc, created_at__lte=day_end_utc)
        day_completed = day_calls.filter(status__in=[CallStatus.COMPLETED, CallStatus.TRANSFERRED]).count()
        day_failed = day_calls.filter(status=CallStatus.FAILED).count()
        day_total = day_calls.count()
        day_success_rate = (day_completed / day_total * 100) if day_total > 0 else 0

        daily_stats.append({
            'date': current_date.strftime('%d.%m'),
            'total_calls': day_total,
            'completed_calls': day_completed,
            'failed_calls': day_failed,
            'success_rate': round(day_success_rate, 1)
        })
        current_date += timedelta(days=1)

    # Team performance breakdown
    team_stats = []
    if region_filter or team_filter:
        # Show individual team stats when filtered
        teams_qs = Team.objects.filter(is_active=True)
        if region_filter:
            teams_qs = teams_qs.filter(region_id=region_filter)
        if team_filter:
            teams_qs = teams_qs.filter(id=team_filter)

        for team in teams_qs[:10]:  # Limit to top 10 teams
            team_calls = callbacks_qs.filter(team=team)
            team_total = team_calls.count()
            team_completed = team_calls.filter(status__in=[CallStatus.COMPLETED, CallStatus.TRANSFERRED]).count()
            team_failed = team_calls.filter(status=CallStatus.FAILED).count()
            team_ratings = ratings_qs.filter(team=team)
            team_avg_rating = team_ratings.aggregate(Avg('rating'))['rating__avg'] or 0

            if team_total > 0:  # Only show teams with calls
                team_stats.append({
                    'team': team,
                    'total_calls': team_total,
                    'completed_calls': team_completed,
                    'failed_calls': team_failed,
                    'success_rate': round((team_completed / team_total * 100), 1),
                    'failure_rate': round((team_failed / team_total * 100), 1),
                    'avg_rating': round(team_avg_rating, 1),
                    'total_ratings': team_ratings.count()
                })
    else:
        # Show region stats when no specific filter
        for region in Region.objects.filter(is_active=True)[:5]:
            region_calls = callbacks_qs.filter(team__region=region)
            region_total = region_calls.count()
            region_completed = region_calls.filter(status__in=[CallStatus.COMPLETED, CallStatus.TRANSFERRED]).count()
            region_failed = region_calls.filter(status=CallStatus.FAILED).count()
            region_ratings = ratings_qs.filter(team__region=region)
            region_avg_rating = region_ratings.aggregate(Avg('rating'))['rating__avg'] or 0

            if region_total > 0:  # Only show regions with calls
                team_stats.append({
                    'team': region,  # Using same template structure
                    'total_calls': region_total,
                    'completed_calls': region_completed,
                    'failed_calls': region_failed,
                    'success_rate': round((region_completed / region_total * 100), 1),
                    'failure_rate': round((region_failed / region_total * 100), 1),
                    'avg_rating': round(region_avg_rating, 1),
                    'total_ratings': region_ratings.count()
                })

    # Sort team stats by total calls
    team_stats.sort(key=lambda x: x['total_calls'], reverse=True)

    # For filter dropdowns
    regions = Region.objects.filter(is_active=True).order_by('name')
    teams = Team.objects.filter(is_active=True).select_related('region').order_by('region__name', 'name')
    if region_filter:
        teams = teams.filter(region_id=region_filter)

    # Calculate period description for display
    if date_from == date_to:
        if date_from == timezone.now().astimezone(TASHKENT_TZ).date():
            period_description = "Сегодня"
        else:
            period_description = f"За {date_from.strftime('%d.%m.%Y')}"
    else:
        period_description = f"С {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')}"

    context = {
        'total_calls': total_calls,
        'completed_calls': completed_calls,
        'failed_calls': failed_calls,
        'no_rating_calls': no_rating_calls,
        'pending_calls': pending_calls,
        'total_ratings': total_ratings,
        'avg_rating': round(avg_rating, 1),
        'success_rate': success_rate,
        'failure_rate': failure_rate,
        'rating_distribution': rating_distribution,
        'recent_calls': recent_calls,
        'daily_stats': daily_stats,
        'team_stats': team_stats,

        # Filter data
        'regions': regions,
        'teams': teams,
        'statuses': CallStatus.choices,
        'current_region': region_filter,
        'current_team': team_filter,
        'current_status': status_filter,
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
        'period_description': period_description,

        # Additional context
        'has_filters': bool(region_filter or team_filter or status_filter or
                            request.GET.get('date_from') or request.GET.get('date_to')),
        'filter_count': sum([bool(region_filter), bool(team_filter), bool(status_filter),
                             bool(request.GET.get('date_from')), bool(request.GET.get('date_to'))]),
    }

    return render(request, 'callbacks/dashboard.html', context)


@has_role([UserRoleChoices.ADMIN, UserRoleChoices.OPERATOR])
def callback_list(request):
    status_filter = request.GET.get('status', '')
    team_filter = request.GET.get('team', '')
    region_filter = request.GET.get('region', '')
    rating_filter = request.GET.get('rating', '')  # New rating filter
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    time_from = request.GET.get('time_from', '')  # New time filter
    time_to = request.GET.get('time_to', '')  # New time filter

    callbacks = CallbackRequest.objects.select_related('team', 'team__region', 'requested_by').prefetch_related(
        'rating').order_by('-created_at')

    # Apply filters
    if status_filter:
        callbacks = callbacks.filter(status=status_filter)

    if team_filter:
        callbacks = callbacks.filter(team_id=team_filter)

    if region_filter:
        callbacks = callbacks.filter(team__region_id=region_filter)

    # Rating filter
    if rating_filter:
        if rating_filter == 'no_rating':
            callbacks = callbacks.filter(rating__isnull=True)
        elif rating_filter == 'high_rating':
            callbacks = callbacks.filter(rating__rating__gte=4)
        elif rating_filter == 'low_rating':
            callbacks = callbacks.filter(rating__rating__lte=2)
        else:
            try:
                rating_value = int(rating_filter)
                if 1 <= rating_value <= 5:
                    callbacks = callbacks.filter(rating__rating=rating_value)
            except ValueError:
                pass

    if search:
        callbacks = callbacks.filter(
            Q(phone_number__icontains=search) |
            Q(team__name__icontains=search) |
            Q(team__region__name__icontains=search)
        )

    # Enhanced date and time filtering
    if date_from:
        from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        if time_from:
            try:
                from_time = datetime.strptime(time_from, '%H:%M').time()
                from_datetime = timezone.make_aware(
                    datetime.combine(from_date, from_time),
                    TASHKENT_TZ
                ).astimezone(pytz.UTC)
            except ValueError:
                from_datetime = timezone.make_aware(
                    datetime.combine(from_date, datetime.min.time()),
                    TASHKENT_TZ
                ).astimezone(pytz.UTC)
        else:
            from_datetime = timezone.make_aware(
                datetime.combine(from_date, datetime.min.time()),
                TASHKENT_TZ
            ).astimezone(pytz.UTC)

        callbacks = callbacks.filter(created_at__gte=from_datetime)

    if date_to:
        to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        if time_to:
            try:
                to_time = datetime.strptime(time_to, '%H:%M').time()
                to_datetime = timezone.make_aware(
                    datetime.combine(to_date, to_time),
                    TASHKENT_TZ
                ).astimezone(pytz.UTC)
            except ValueError:
                to_datetime = timezone.make_aware(
                    datetime.combine(to_date, datetime.max.time()),
                    TASHKENT_TZ
                ).astimezone(pytz.UTC)
        else:
            to_datetime = timezone.make_aware(
                datetime.combine(to_date, datetime.max.time()),
                TASHKENT_TZ
            ).astimezone(pytz.UTC)

        callbacks = callbacks.filter(created_at__lte=to_datetime)

    # For filters
    regions = Region.objects.filter(is_active=True).order_by('name')
    teams = Team.objects.filter(is_active=True).select_related('region').order_by('region__name', 'name')
    if region_filter:
        teams = teams.filter(region_id=region_filter)

    statuses = CallStatus.choices

    # Rating filter options
    rating_options = [
        ('', 'Все оценки'),
        ('no_rating', 'Без оценки'),
        ('high_rating', 'Высокие (4-5★)'),
        ('low_rating', 'Низкие (1-2★)'),
        ('1', '1 звезда'),
        ('2', '2 звезды'),
        ('3', '3 звезды'),
        ('4', '4 звезды'),
        ('5', '5 звезд'),
    ]

    # Pagination-like limit
    total_count = callbacks.count()
    callbacks = callbacks[:100]  # Increased limit

    context = {
        'callbacks': callbacks,
        'regions': regions,
        'teams': teams,
        'statuses': statuses,
        'rating_options': rating_options,
        'current_status': status_filter,
        'current_team': team_filter,
        'current_region': region_filter,
        'current_rating': rating_filter,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
        'time_from': time_from,
        'time_to': time_to,
        'total_count': total_count,
        'showing_count': len(callbacks),
    }

    return render(request, 'callbacks/list.html', context)


@has_role([UserRoleChoices.ADMIN, UserRoleChoices.OPERATOR])
def callback_create(request):
    if request.method == 'POST':
        form = CallbackRequestForm(request.POST)
        if form.is_valid():
            callback = form.save(commit=False)
            callback.requested_by = request.user
            callback.save()

            # Process call asynchronously
            process_callback_call.delay(callback.id)

            messages.success(request, get_message('callback_created', phone_number=callback.phone_number))
            return redirect('callbacks:list')
        else:
            messages.error(request, get_message('invalid_data'))
    else:
        form = CallbackRequestForm()

    # Quick stats for the form page (today only)
    today = timezone.now().astimezone(TASHKENT_TZ).date()
    today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()), TASHKENT_TZ).astimezone(pytz.UTC)
    today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()), TASHKENT_TZ).astimezone(pytz.UTC)

    today_calls = CallbackRequest.objects.filter(created_at__gte=today_start, created_at__lte=today_end).count()
    today_completed = CallbackRequest.objects.filter(
        created_at__gte=today_start,
        created_at__lte=today_end,
        status__in=[CallStatus.COMPLETED, CallStatus.TRANSFERRED]
    ).count()
    today_failed = CallbackRequest.objects.filter(
        created_at__gte=today_start,
        created_at__lte=today_end,
        status=CallStatus.FAILED
    ).count()
    today_ratings = Rating.objects.filter(
        timestamp__gte=today_start,
        timestamp__lte=today_end
    ).count()

    # Calculate success rate for the form page
    today_success_rate = 0
    if today_calls > 0:
        today_success_rate = round((today_completed / today_calls) * 100, 0)

    # Get regions and teams for the enhanced form
    regions = Region.objects.filter(is_active=True).order_by('name')
    teams = Team.objects.filter(is_active=True).select_related('region').order_by('region__name', 'name')

    context = {
        'form': form,
        'title': 'Создать экстренный вызов',
        'today_calls': today_calls,
        'today_completed': today_completed,
        'today_failed': today_failed,
        'today_ratings': today_ratings,
        'today_success_rate': today_success_rate,
        'regions': regions,
        'teams': teams,
    }

    return render(request, 'callbacks/form.html', context)


# Rest of the views remain the same with updated Russian messages...
@has_role([UserRoleChoices.ADMIN, UserRoleChoices.OPERATOR])
def callback_detail(request, pk):
    callback = get_object_or_404(CallbackRequest, pk=pk)

    # Get rating if exists
    rating = None
    try:
        rating = callback.rating
    except:
        pass

    context = {
        'callback': callback,
        'rating': rating,
    }

    return render(request, 'callbacks/detail.html', context)


@admin_required
def ratings_list(request):
    team_filter = request.GET.get('team', '')
    region_filter = request.GET.get('region', '')
    rating_filter = request.GET.get('rating', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    ratings = Rating.objects.select_related('team', 'team__region', 'callback_request').order_by('-timestamp')

    # Apply filters
    if team_filter:
        ratings = ratings.filter(team_id=team_filter)

    if region_filter:
        ratings = ratings.filter(team__region_id=region_filter)

    if rating_filter:
        ratings = ratings.filter(rating=rating_filter)

    if date_from:
        from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        from_datetime = timezone.make_aware(
            datetime.combine(from_date, datetime.min.time()),
            TASHKENT_TZ
        ).astimezone(pytz.UTC)
        ratings = ratings.filter(timestamp__gte=from_datetime)

    if date_to:
        to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        to_datetime = timezone.make_aware(
            datetime.combine(to_date, datetime.max.time()),
            TASHKENT_TZ
        ).astimezone(pytz.UTC)
        ratings = ratings.filter(timestamp__lte=to_datetime)

    # Calculate statistics for filtered ratings
    total_ratings = ratings.count()
    avg_rating = ratings.aggregate(Avg('rating'))['rating__avg'] or 0

    # Rating distribution for filtered results
    rating_stats = {}
    for i in range(1, 6):
        count = ratings.filter(rating=i).count()
        percentage = (count / total_ratings * 100) if total_ratings > 0 else 0
        rating_stats[i] = {
            'count': count,
            'percentage': round(percentage, 1)
        }

    # Good ratings (4+ stars)
    good_ratings = ratings.filter(rating__gte=4).count()
    good_percentage = (good_ratings / total_ratings * 100) if total_ratings > 0 else 0

    # For filters
    regions = Region.objects.filter(is_active=True).order_by('name')
    teams = Team.objects.filter(is_active=True).select_related('region').order_by('region__name', 'name')
    if region_filter:
        teams = teams.filter(region_id=region_filter)

    # Limit for performance
    ratings_list = ratings[:100]

    context = {
        'ratings': ratings_list,
        'regions': regions,
        'teams': teams,
        'current_team': team_filter,
        'current_region': region_filter,
        'current_rating': rating_filter,
        'date_from': date_from,
        'date_to': date_to,
        'total_ratings': total_ratings,
        'avg_rating': round(avg_rating, 1),
        'rating_stats': rating_stats,
        'good_percentage': round(good_percentage, 1),
        'showing_count': len(ratings_list),
    }

    return render(request, 'callbacks/ratings.html', context)


# AJAX endpoint for getting teams by region
@has_role([UserRoleChoices.ADMIN, UserRoleChoices.OPERATOR])
def get_teams_by_region(request):
    region_id = request.GET.get('region_id')
    teams = []

    if region_id:
        teams_qs = Team.objects.filter(region_id=region_id, is_active=True).order_by('name')
        teams = [{'id': team.id, 'name': team.name} for team in teams_qs]

    return JsonResponse({'teams': teams})


@csrf_exempt
@require_http_methods(["POST"])
def api_callback_create(request):
    try:
        # Parse JSON data
        data = json.loads(request.body)
        phone_number = data.get('phone_number')

        if not phone_number:
            return JsonResponse({'error': get_message('invalid_phone')}, status=400)

        # Get first user for requested_by
        from users.models import User
        user = User.objects.first()
        if not user:
            return JsonResponse({'error': get_message('system_error')}, status=500)

        # Get random active team
        teams = list(Team.objects.filter(is_active=True))
        if not teams:
            return JsonResponse({'error': get_message('no_active_teams')}, status=500)

        team = random.choice(teams)

        # Create callback request
        callback = CallbackRequest.objects.create(
            phone_number=phone_number,
            team=team,
            requested_by=user
        )

        # Process call asynchronously
        process_callback_call.delay(callback.id)

        return JsonResponse({
            'success': True,
            'callback_id': callback.id,
            'phone_number': callback.phone_number,
            'team': callback.team.name,
            'region': callback.team.region.name if hasattr(callback.team, 'region') and callback.team.region else None,
            'status': callback.status,
            'message': get_message('callback_created', phone_number=callback.phone_number)
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': get_message('invalid_data')}, status=400)
    except Exception as e:
        return JsonResponse({'error': get_message('system_error')}, status=500)


@admin_required
def export_excel(request):
    """Export dashboard data to Excel format - includes ALL teams/regions"""
    try:
        # Get filter parameters (same as dashboard)
        region_filter = request.GET.get('region', '')
        team_filter = request.GET.get('team', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        status_filter = request.GET.get('status', '')

        # Default date range (today if no filters)
        if not date_from and not date_to:
            today = timezone.now().astimezone(TASHKENT_TZ).date()
            date_from = today
            date_to = today
        else:
            if date_from:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            else:
                date_from = timezone.now().astimezone(TASHKENT_TZ).date() - timedelta(days=30)

            if date_to:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            else:
                date_to = timezone.now().astimezone(TASHKENT_TZ).date()

        # Convert dates to UTC for database queries
        start_utc = timezone.make_aware(datetime.combine(date_from, datetime.min.time()), TASHKENT_TZ).astimezone(
            pytz.UTC)
        end_utc = timezone.make_aware(datetime.combine(date_to, datetime.max.time()), TASHKENT_TZ).astimezone(pytz.UTC)

        # Base querysets with date filtering
        callbacks_qs = CallbackRequest.objects.filter(
            created_at__gte=start_utc,
            created_at__lte=end_utc
        )

        ratings_qs = Rating.objects.filter(
            timestamp__gte=start_utc,
            timestamp__lte=end_utc
        )

        # Apply region filter to querysets only (not to teams selection)
        if region_filter:
            callbacks_qs = callbacks_qs.filter(team__region_id=region_filter)
            ratings_qs = ratings_qs.filter(team__region_id=region_filter)

        # Apply team filter to querysets only
        if team_filter:
            callbacks_qs = callbacks_qs.filter(team_id=team_filter)
            ratings_qs = ratings_qs.filter(team_id=team_filter)

        # Apply status filter
        if status_filter:
            callbacks_qs = callbacks_qs.filter(status=status_filter)

        # Get ALL active teams (filtered by region/team if specified, but include all active ones)
        teams_query = Team.objects.filter(is_active=True).select_related('region')

        # Only filter teams if specific filters are applied
        if region_filter:
            teams_query = teams_query.filter(region_id=region_filter)
        if team_filter:
            teams_query = teams_query.filter(id=team_filter)

        teams_query = teams_query.order_by('region__name', 'name')

        # Collect data for Excel - INCLUDE ALL TEAMS even with 0 calls
        excel_data = []

        for team in teams_query:
            # Get team statistics
            team_callbacks = callbacks_qs.filter(team=team)
            team_ratings = ratings_qs.filter(team=team)

            total_calls = team_callbacks.count()
            success_calls = team_callbacks.filter(
                status__in=[CallStatus.COMPLETED, CallStatus.TRANSFERRED]
            ).count()
            failed_calls = team_callbacks.filter(status=CallStatus.FAILED).count()

            # Rating counts
            rating_counts = {i: 0 for i in range(1, 6)}
            for rating in team_ratings:
                if 1 <= rating.rating <= 5:
                    rating_counts[rating.rating] += 1

            # Average rating
            avg_rating = team_ratings.aggregate(Avg('rating'))['rating__avg'] or 0

            # INCLUDE ALL TEAMS - even those with 0 calls
            excel_data.append({
                'region_team': f"{team.region.name} - {team.name}",
                'total_calls': total_calls,
                'success_calls': success_calls,
                'failed_calls': failed_calls,
                'rating_5': rating_counts[5],
                'rating_4': rating_counts[4],
                'rating_3': rating_counts[3],
                'rating_2': rating_counts[2],
                'rating_1': rating_counts[1],
                'avg_rating': round(avg_rating, 2)
            })

        # Create Excel file
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Set sheet title
        period_str = f"с {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')}"
        worksheet.title = f"Отчет {period_str}"

        # Headers
        headers = [
            '№',
            'Регион - Бригада',
            'Всего вызовов',
            'Успешных',
            'Неудачных',
            '5 звезд',
            '4 звезды',
            '3 звезды',
            '2 звезды',
            '1 звезда',
            'Средняя оценка'
        ]

        # Add title row
        worksheet.merge_cells('A1:K1')
        title_cell = worksheet['A1']
        title_cell.value = f"Отчет по обратным вызовам {period_str}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

        # Add headers in row 3
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

        # Add data - ALL teams are included now
        for row_num, data in enumerate(excel_data, start=4):
            worksheet.cell(row=row_num, column=1, value=row_num - 3)  # Numbering
            worksheet.cell(row=row_num, column=2, value=data['region_team'])
            worksheet.cell(row=row_num, column=3, value=data['total_calls'])
            worksheet.cell(row=row_num, column=4, value=data['success_calls'])
            worksheet.cell(row=row_num, column=5, value=data['failed_calls'])
            worksheet.cell(row=row_num, column=6, value=data['rating_5'])
            worksheet.cell(row=row_num, column=7, value=data['rating_4'])
            worksheet.cell(row=row_num, column=8, value=data['rating_3'])
            worksheet.cell(row=row_num, column=9, value=data['rating_2'])
            worksheet.cell(row=row_num, column=10, value=data['rating_1'])
            worksheet.cell(row=row_num, column=11, value=data['avg_rating'])

        # Style the data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply borders and alignment to all cells with data
        for row in worksheet.iter_rows(min_row=3, max_row=len(excel_data) + 3, min_col=1, max_col=11):
            for cell in row:
                cell.border = thin_border
                if cell.column == 2:  # Region-Team column
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        column_widths = {
            'A': 5,  # №
            'B': 30,  # Регион - Бригада
            'C': 12,  # Всего вызовов
            'D': 12,  # Успешных
            'E': 12,  # Неудачных
            'F': 10,  # 5 звезд
            'G': 10,  # 4 звезды
            'H': 10,  # 3 звезды
            'I': 10,  # 2 звезды
            'J': 10,  # 1 звезда
            'K': 15,  # Средняя оценка
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        # Add totals row - now includes all teams
        if excel_data:
            total_row = len(excel_data) + 5

            # Add totals
            worksheet.cell(row=total_row, column=1, value="ИТОГО:")
            worksheet.cell(row=total_row, column=1).font = Font(bold=True)

            # Calculate totals
            total_calls_sum = sum(data['total_calls'] for data in excel_data)
            total_success_sum = sum(data['success_calls'] for data in excel_data)
            total_failed_sum = sum(data['failed_calls'] for data in excel_data)
            total_rating_5 = sum(data['rating_5'] for data in excel_data)
            total_rating_4 = sum(data['rating_4'] for data in excel_data)
            total_rating_3 = sum(data['rating_3'] for data in excel_data)
            total_rating_2 = sum(data['rating_2'] for data in excel_data)
            total_rating_1 = sum(data['rating_1'] for data in excel_data)

            # Calculate overall average rating
            total_ratings = sum(data['rating_5'] * 5 + data['rating_4'] * 4 + data['rating_3'] * 3 +
                                data['rating_2'] * 2 + data['rating_1'] * 1 for data in excel_data)
            total_rating_count = sum(data['rating_5'] + data['rating_4'] + data['rating_3'] +
                                     data['rating_2'] + data['rating_1'] for data in excel_data)
            overall_avg = round(total_ratings / total_rating_count, 2) if total_rating_count > 0 else 0

            worksheet.cell(row=total_row, column=3, value=total_calls_sum)
            worksheet.cell(row=total_row, column=4, value=total_success_sum)
            worksheet.cell(row=total_row, column=5, value=total_failed_sum)
            worksheet.cell(row=total_row, column=6, value=total_rating_5)
            worksheet.cell(row=total_row, column=7, value=total_rating_4)
            worksheet.cell(row=total_row, column=8, value=total_rating_3)
            worksheet.cell(row=total_row, column=9, value=total_rating_2)
            worksheet.cell(row=total_row, column=10, value=total_rating_1)
            worksheet.cell(row=total_row, column=11, value=overall_avg)

            # Style totals row
            for col in range(1, 12):
                cell = worksheet.cell(row=total_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                cell.border = thin_border

        # Prepare response
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # Create filename
        filename = f"otchet_vyzovy_{date_from.strftime('%d_%m_%Y')}-{date_to.strftime('%d_%m_%Y')}.xlsx"

        # Create HTTP response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        # Log the error
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Excel export error: {str(e)}")

        # Return error response
        return HttpResponse(
            f'Ошибка при экспорте: {str(e)}',
            status=500,
            content_type='text/plain; charset=utf-8'
        )